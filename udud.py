import os
from openpyxl import load_workbook, Workbook
from tkinter import Tk, Label, Entry, Button, Listbox, END, Scrollbar
import pandas as pd

rootError=Tk()
rootError.title("Вывод ответов сервера")
rootError.geometry("400x300+1100+50")

listboxerr = Listbox(rootError, width=60, height=20)
listboxerr.grid(row=0, column=0, columnspan=10)
listboxerr.config(bg="#FFF0F5")
scrollerr = Scrollbar(rootError)
scrollerr.grid(row=0, column=10, sticky="ns")
listboxerr.config(yscrollcommand=scrollerr.set)
scrollerr.config(command=listboxerr.yview)



def load_workbook_safe(file_path):
    """Безопасная загрузка рабочей книги."""
    try:
        return load_workbook(file_path)
    except FileNotFoundError:
        listboxerr.insert(END,f"Файл {file_path} не найден.")
        return None

def big_del_ser_row(file_path, row_name, the_value, flag):
    """Удаление строки из Excel-файла"""
    wb = load_workbook_safe(file_path)    
    sheet = wb.active

    # Находим индекс строки
    row_index = None
    tmp=[[]]
    for i in range(sheet.max_row, 0, -1):
        cell = sheet.cell(row=i, column=the_value)
        if cell.value == row_name:
            row_index=i
            if flag==1:
                sheet.delete_rows(i, amount=1)  
            else:
                arr=[]
                for j in range (1, sheet.max_column + 1):
                    arr.append(sheet.cell(row=i, column=j).value)
                tmp.append(arr)
    for k in range (len(tmp)-1, 0, -1):
        listboxerr.insert(END, tmp[k])

    if row_index is None:
        listboxerr.insert(END, f"Строка '{row_name}' не найден.")
        return

    wb.save(file_path)
    if flag==1:
        listboxerr.insert(END, f"Строки со значением '{row_name}' удалены.")
    else:
        listboxerr.insert(END, f"Строки со значением '{row_name}' найдены.")


def append_row(file_path, row_values):
    """Добавление новой строки"""
    wb = load_workbook_safe(file_path)
    if not wb:
        return

    sheet = wb.active
    sheet.append(row_values)
    wb.save(file_path)
    listboxerr.insert(END,f"Новая строка добавлена: {row_values}")

def display_table(file_path):
    """Вывод данных из Excel-файла для визуализации."""
    wb = load_workbook_safe(file_path)
    if not wb:
        return

    sheet = wb.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return data

def refresh_listbox(file_path, listbox):
    """Обновляет содержимое Listbox."""
    listbox.delete(0, END)
    data = display_table(file_path)
    if data:
        for row in data:
            listbox.insert(END, row)

def sorting(file_path):
    "Сортировка для добавления и удаления"
    wb=load_workbook_safe(file_path)
    sheet=wb.active
    rows=list(sheet.iter_rows(values_only=True))[1:]
    sorted_rows=sorted(rows, key=lambda x: int(x[0]))
    sheet.delete_rows(2, sheet.max_row)
    for row in sorted_rows:
        sheet.append(row)
    wb.save(file_path)

def binary_search(file_path, value):
    """Бинарный поиск"""
    wb=load_workbook_safe(file_path)
    sheet=wb.active
    rows=list(sheet.iter_rows(values_only=True))[1:]
    left, right=0, len(rows)-1
    while left<=right:
        mid=(left+right)//2
        if int(rows[mid][0])==value:
            return mid+1
        elif value>int(rows[mid][0]):
            left=mid+1
        elif value<int(rows[mid][0]):
            right=mid-1
    return -1

def gui_interface():
    """Графический интерфейс для взаимодействия с файлом Excel."""
    file_path = "example.xlsx"
    if not os.path.exists(file_path):
        workbook= Workbook()
        workbook.save(file_path)
    wb = load_workbook_safe(file_path)
    sheet=wb.active
    # Создание окна
    root = Tk()
    root.title("Управление Excel-файлом")
    root.geometry("1000x460+50+50")
    

    # Элементы интерфейса
    listbox = Listbox(root, width=200, height=20)
    listbox.grid(row=0, column=0, columnspan=10)
    listbox.config(bg="#FFF0F5")

    scrollbar = Scrollbar(root)
    scrollbar.grid(row=0, column=10, sticky="ns")
    listbox.config(yscrollcommand=scrollbar.set)
    scrollbar.config(command=listbox.yview)
    
    num_columns = 8
    for col in range(num_columns):
        root.grid_columnconfigure(col, weight=1, uniform="equal")

    if sheet.cell(row=1, column=1).value is None:
        append_row(file_path, ["ID", "Name", "Type", "Age", "Weight"])
        refresh_listbox(file_path, listbox)

    """Поля для ввода данных"""
    Label(root, text="ID:").grid(row=1, column=0)
    id_entry = Entry(root)
    id_entry.grid(row=1, column=1)

    Label(root, text="Name:").grid(row=1, column=2)
    name_entry = Entry(root)
    name_entry.grid(row=1, column=3)

    Label(root, text="Type:").grid(row=1, column=4)
    type_entry=Entry(root)
    type_entry.grid(row=1, column=5)

    Label(root, text="Age:").grid(row=1, column=6)
    age_entry=Entry(root)
    age_entry.grid(row=1, column=7)

    Label(root, text="Weight:").grid(row=1, column=8)
    weight_entry=Entry(root)
    weight_entry.grid(row=1, column=9)
    
    id_label=Label(root, text="Введите ID:")
    id_label.grid(row=4, column=0)

    name_label=Label(root, text="Введите имя:")
    name_label.grid(row=4, column=0)

    type_label=Label(root, text="Введите породу:")
    type_label.grid(row=4, column=0)

    age_label=Label(root, text="Введите возраст:")
    age_label.grid(row=4, column=0)

    weight_label=Label(root, text="Введите наличие дома:")
    weight_label.grid(row=4, column=0)

    id_del_ser = Entry(root)
    id_del_ser.grid(row=4, column=1)

    name_del_ser = Entry(root)
    name_del_ser.grid(row=4, column=1)

    type_del_ser = Entry(root)
    type_del_ser.grid(row=4, column=1)

    age_del_ser = Entry(root)
    age_del_ser.grid(row=4, column=1)

    weight_del_ser = Entry(root)
    weight_del_ser.grid(row=4, column=1)
    """Скрыть кнопки и поля ввода"""
    def hide():
        id_label.grid_remove()
        id_del_ser.grid_remove()

        name_label.grid_remove()
        name_del_ser.grid_remove()

        type_label.grid_remove()
        type_del_ser.grid_remove()

        age_label.grid_remove()
        age_del_ser.grid_remove()

        weight_label.grid_remove()
        weight_del_ser.grid_remove()
    hide()
    def hide_button():
        search_id_button.grid_remove()
        search_name_button.grid_remove()
        search_type_button.grid_remove()
        search_age_button.grid_remove()
        search_weight_button.grid_remove()

        delete_id_button.grid_remove()
        delete_name_button.grid_remove()
        delete_type_button.grid_remove()
        delete_age_button.grid_remove()
        delete_weight_button.grid_remove()

    """Удаление/поиск по ID"""
    def del_ser_by_id(flag):
        if flag==2:
            id = id_entry.get()
            return binary_search(file_path, int(id)) + 1
        hide()
        id_label.grid()
        id_del_ser.grid()

        def del_ser_row():
            id=id_del_ser.get()
            i=binary_search(file_path, int(id))
            print(i)
            if i==-1:
                listboxerr.insert(END, "Элемента с таким индексом не существует")
                refresh_listbox(file_path, listbox)
                return -1
            if flag==1:
                sheet.delete_rows(i+1, amount=1) 
                wb.save(file_path)
                listboxerr.insert(END, f"Строка с индексом {id} удалена")
            elif flag==0:
                tmp=[[]]
                arr=[]
                for j in range (1, sheet.max_column + 1):
                    arr.append(sheet.cell(row=i+1, column=j).value)
                tmp.append(arr)
                listboxerr.insert(END, arr) 
            refresh_listbox(file_path, listbox)
            hide()
            hide_button()
            del_sers.grid_remove()
            id_del_ser.delete(0, END)
        del_sers=Button(root, text="Выполнить", command=del_ser_row)
        del_sers.grid(row=5, column=1)

    """Удаление/поиск по имени"""
    def del_ser_by_name(flag):
        hide()
        name_label.grid()
        name_del_ser.grid()
    
        def del_ser_row():
            key=name_del_ser.get()
            big_del_ser_row(file_path, key, 2, flag)
            refresh_listbox(file_path, listbox)
            hide()
            hide_button()
            del_sers.grid_remove()
            name_del_ser.delete(0, END)
        del_sers=Button(root, text="Выполнить", command=del_ser_row)
        del_sers.grid(row=5, column=1)

    """Удаление/поиск по породе"""
    def del_ser_by_type(flag):
        hide()
        type_label.grid()
        type_del_ser.grid()

        def del_ser_row():
            key=type_del_ser.get()
            big_del_ser_row(file_path, key, 3, flag)
            refresh_listbox(file_path, listbox)
            hide()
            hide_button()
            del_sers.grid_remove()
            type_del_ser.delete(0, END)
        del_sers=Button(root, text="Выполнить", command=del_ser_row)
        del_sers.grid(row=5, column=1)

    """Удаление/поиск по возрасту"""
    def del_ser_by_age(flag):
        hide()
        age_label.grid()
        age_del_ser.grid()

        def del_ser_row():
            key=age_del_ser.get()
            big_del_ser_row(file_path, key, 4, flag)
            refresh_listbox(file_path, listbox)
            hide()
            hide_button()
            del_sers.grid_remove()
            age_del_ser.delete(0, END)
        del_sers=Button(root, text="Выполнить", command=del_ser_row)
        del_sers.grid(row=5, column=1)

    """Удаление/поиск по весу"""
    def del_ser_by_weight(flag):
        hide()
        weight_label.grid()
        weight_del_ser.grid()

        def del_ser_row():
            key=weight_del_ser.get()
            big_del_ser_row(file_path, key, 5, flag)
            refresh_listbox(file_path, listbox)
            hide()
            hide_button()
            del_sers.grid_remove()
            weight_del_ser.delete(0, END)
        del_sers=Button(root, text="Выполнить", command=del_ser_row)
        del_sers.grid(row=5, column=1)

    """Кнопки удаления появляются"""
    def delete():
        delete_id_button.grid(row=3,column=0)
        delete_name_button.grid(row=3,column=1)
        delete_type_button.grid(row=3,column=2)
        delete_age_button.grid(row=3,column=3)
        delete_weight_button.grid(row=3,column=4)

    """Проверка введённого возраста и веса"""
    def check_age(age):
        try:
            x=int(age)
        except ValueError:
            return 0
    def check_weight(weight):
        try:
            y=float(weight)
        except ValueError:
            return 0

    """"Добавление строки"""
    def add_row():
        wb = load_workbook_safe(file_path)
        sheet=wb.active
        for i in range (1, sheet.max_row+1):
            if (sheet.cell(row=i, column=1).value == id_entry.get()):
                listboxerr.insert(END, "Объект с таким iD уже существует")
                return
        if (age_entry.get()) and (check_age(age_entry.get())==0):
            listboxerr.insert(END, "Возраст должен быть целым числом")
            return
        if (weight_entry.get()) and (check_weight(weight_entry.get())==0):
            listboxerr.insert(END, "Вес должен быть дробным числом")
            return
        new_values = [id_entry.get(), name_entry.get(), type_entry.get(), age_entry.get(), weight_entry.get()]
        append_row(file_path, new_values)
        sorting(file_path)

        refresh_listbox(file_path, listbox)
        id_entry.delete(0, END)
        name_entry.delete(0, END)
        type_entry.delete(0, END)
        age_entry.delete(0,END)
        weight_entry.delete(0,END)

    """Полная очистка БД"""
    def delete_all():
        wb = load_workbook_safe(file_path)
        if not wb:
            return
        sheet=wb.active
        sheet.delete_cols(1, sheet.max_column)
        wb.save(file_path)
        refresh_listbox(file_path, listbox)
        listboxerr.insert(END, "Файл полностью очищен")
        if sheet.cell(row=1, column=1).value is None:
            append_row(file_path, ["ID", "Name", "Type", "Age", "Weight"])
            refresh_listbox(file_path, listbox)

    """"Удаление БД"""
    def remove_file():
        os.remove(file_path)
        root.destroy()
        rootError.destroy()

    """Редактирование строки"""
    def edit():
        wb=load_workbook_safe(file_path)
        sheet=wb.active
        new_values=[id_entry.get(), name_entry.get(), type_entry.get(), age_entry.get(), weight_entry.get()]
        id=del_ser_by_id(2)
        if id==0:
            listboxerr.insert(END, f"Строка с индексом '{id_entry.get()}' не найден.")
            return
        for i in range (1, sheet.max_column + 1):
            sheet.cell(row=id, column=i).value=new_values[i-1]

        wb.save(file_path)
        listboxerr.insert(END, f"Строка с индексом '{id_entry.get()}' изменена.")
        refresh_listbox(file_path, listbox)

    """Создание бэкапа"""
    def create_backup():
        copy_path = "copy.xlsx"
        if not os.path.exists(copy_path):
            copy_book= Workbook()
            copy_book.save(copy_path)
        wb=load_workbook_safe(file_path)
        wb.save(copy_path)
        listboxerr.insert(END, "Backup успешно создался")

    """Загрузка последней версии"""
    def reload_file():
        copy_path="copy.xlsx"
        if not os.path.exists(file_path):
            wb=Workbook()
            wb.save(file_path)
        copy_book=load_workbook_safe(copy_path)
        copy_book.save(file_path)
        refresh_listbox(file_path, listbox)
        listboxerr.insert(END, "Файл успешно восстановился из backup-файла")

    """Экспорт в файл формата CSV"""
    def export_csv():
        wb = pd.read_excel ("example.xlsx") 
        csv_path="example.csv"
        wb.to_csv("Example.csv", index=False, sep=';')
        listboxerr.insert(END, "Файл успешно экспортирован в файл в формате CSV")
    
    """Появление кнопок для поиска"""
    def search():
        search_id_button.grid(row=3,column=0)
        search_name_button.grid(row=3,column=1)
        search_type_button.grid(row=3,column=2)
        search_age_button.grid(row=3,column=3)
        search_weight_button.grid(row=3,column=4)

    """Все изначальные кнопки"""
    Button(root, text="Удалить", command=delete, bg="#98FB98").grid(row=2, column=0, sticky="ew", padx=3, pady=5)
    search_id_button=Button(root, text="По ID", command=lambda: del_ser_by_id(0), bg="#FF69B4")
    search_name_button=Button(root, text="По имени", command=lambda : del_ser_by_name(0), bg="#FF69B4")
    search_type_button=Button(root, text="По породе", command=lambda : del_ser_by_type(0), bg="#FF69B4")
    search_age_button=Button(root, text="По возрасту", command=lambda : del_ser_by_age(0), bg="#FF69B4")
    search_weight_button=Button(root, text="По весу", command=lambda : del_ser_by_weight(0), bg="#FF69B4")

    delete_id_button = Button(root, text="По ID", command=lambda:del_ser_by_id(1), bg="#66CDAA")
    delete_name_button = Button(root, text="По имени", command=lambda: del_ser_by_name(1), bg="#66CDAA")
    delete_type_button = Button(root, text="По породе", command=lambda: del_ser_by_type(1), bg="#66CDAA")
    delete_age_button = Button(root, text="По возрасту", command=lambda: del_ser_by_age(1), bg="#66CDAA")
    delete_weight_button = Button(root, text="По весу", command=lambda: del_ser_by_weight(1), bg="#66CDAA")

    Button(root, text="Добавить", command=add_row, bg="#FA8072").grid(row=2, column=1, sticky="ew", padx=3, pady=5)
    Button(root, text="Найти", command=search, bg="#FFC0CB").grid(row=2, column=2, sticky="ew", padx=3, pady=5)
    Button(root, text="Очистить всё", command=delete_all, bg="#87CEEB").grid(row=2, column=3, sticky="ew", padx=3, pady=5)
    Button(root, text="Удалить БД", command=remove_file, bg="#E6E6FA").grid(row=2, column=4, sticky="ew", padx=3, pady=5)
    Button(root, text="Изменить БД", command=edit, bg="#FFDEAD").grid(row=2, column=5, sticky="ew", padx=3, pady=5)
    Button (root, text="Сделать backup", command=create_backup, bg="#9ACD32").grid(row=2, column=6, sticky="ew", padx=3, pady=5)
    Button(root, text="Вернуть backup", command=reload_file, bg="#FFD700").grid(row=2, column=7, sticky="ew", padx=3, pady=5)
    Button(root, text="Экспорт БД в CSV", command=export_csv, bg="#CD5C5C").grid(row=2, column=8, sticky="ew", padx=3, pady=5)


    """Инициализация таблицы"""
    refresh_listbox(file_path, listbox)

    root.mainloop()
    rootError.mainloop()

"""Запуск GUI"""
if __name__ == "__main__":
    gui_interface()
